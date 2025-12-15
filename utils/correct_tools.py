import os
import base64
import mimetypes
import zipfile
from typing import List, Union, Optional
import requests
import io
from PIL import Image
import json
import yt_dlp
from html.parser import HTMLParser
import xml.etree.ElementTree as ET
import nltk
import openpyxl

import tempfile

nltk_data_path = tempfile.mkdtemp()


if nltk_data_path not in nltk.data.path:
    nltk.data.path.append(nltk_data_path)

try:
    nltk.data.find('tokenizers/punkt')
except LookupError:
    print(f"NLTK 'punkt' tokenizer not found. Downloading to {nltk_data_path}...")
    nltk.download('punkt', download_dir=nltk_data_path)

# LangChain and document loaders
from langchain_community.document_loaders import (
    UnstructuredWordDocumentLoader,
    UnstructuredPowerPointLoader,
    UnstructuredPDFLoader
)
from langchain_community.tools.file_management import WriteFileTool, ListDirectoryTool

# PDF specific libraries
from pypdf import PdfReader, PdfWriter

# Smol-Agents framework and OpenAI client
from smolagents import Tool, tool
from smolagents.default_tools import WikipediaSearchTool
from openai import OpenAI
from dotenv import load_dotenv

load_dotenv()

stt_key = os.getenv("STT_KEY")
stt_url = os.getenv("STT_URL")
stt_client = None
if stt_key:
    try:
        stt_client = OpenAI(api_key=stt_key, base_url=stt_url)
        print("STT (OpenAI Whisper) client initialized successfully.")
    except Exception as e:
        print(f"Warning: Failed to initialize STT client. Please check STT_KEY/STT_URL. Error: {e}")
else:
    print("Warning: STT_KEY is not set. The speech_to_text tool will be unavailable.")

vlm_key = os.getenv("VLM_KEY")
vlm_url = os.getenv("VLM_URL")
vlm_model = os.getenv("VLM_MODEL")
vlm_client = None
if vlm_url and vlm_model:
    try:
        vlm_client = OpenAI(api_key=vlm_key, base_url=vlm_url)
        print(f"VLM (Image Analysis) client initialized successfully, pointing to: {vlm_url}")
        print(f"Default VLM model for image analysis will be: '{vlm_model}'")
    except Exception as e:
        print(f"Error: Failed to initialize VLM client. Error: {e}")
else:
    print("Warning: VLM_URL or VLM_MODEL is not set. The analyze_image tool will be unavailable.")

video_key = os.getenv("VIDEO_KEY")
video_url = os.getenv("VIDEO_URL")
video_model = os.getenv("VIDEO_MODEL")
video_client = None
if video_url and video_key and video_model:
    try:
        video_client = OpenAI(api_key=video_key, base_url=video_url)
        print(f"Video Analysis client initialized successfully, pointing to: {video_url}")
        print(f"Default model for video analysis will be: '{video_model}'")
    except Exception as e:
        print(f"Error: Failed to initialize Video Analysis client. Error: {e}")
else:
    print("Warning: VIDEO_URL, VIDEO_KEY, or VIDEO_MODEL is not set. The analyze_video tool will be unavailable.")

doc_key = os.getenv("DOC_KEY")
doc_url = os.getenv("DOC_URL")
doc_model = os.getenv("DOC_MODEL")
doc_client = None
if doc_url and doc_key and doc_model:
    try:
        doc_client = OpenAI(api_key=doc_key, base_url=doc_url)
        print(f"Document Analysis client initialized successfully, pointing to: {doc_url}")
        print(f"Default model for document analysis will be: '{doc_model}'")
    except Exception as e:
        print(f"Error: Failed to initialize Document Analysis client. Error: {e}")
else:
    print(
        "Warning: DOC_URL, DOC_KEY, or DOC_MODEL is not set. The ask_question_about_complex_document tool will be unavailable.")

MAX_PAGES_TO_PROCESS = 10

def _encode_pillow_image_to_base64(image: Image.Image) -> str:
    buffered = io.BytesIO()
    if image.mode not in ('RGB', 'L'):
        image = image.convert('RGB')
    image.save(buffered, format="PNG")
    img_bytes = buffered.getvalue()
    base64_str = base64.b64encode(img_bytes).decode('utf-8')
    return f"data:image/png;base64,{base64_str}"


def _get_file_mime_type(file_path: str) -> str:
    mime_type, _ = mimetypes.guess_type(file_path)
    return mime_type or "application/octet-stream"


@tool
def speech_to_text(audio_file_path: str) -> str:
    """
    Transcribes an audio file into text using the OpenAI Whisper API.
    It returns a formatted string containing the source filename and the transcribed text, or an error message starting with 'Error:' upon failure.

    Args:
        audio_file_path: The local file path to the audio file (e.g., .mp3, .wav, .m4a) to be transcribed.
    """
    if not stt_client:
        return "Error: STT client is not initialized. Cannot call the Whisper API."
    if not os.path.exists(audio_file_path):
        return f"Error: Audio file not found at path: {audio_file_path}"

    print(f"Transcribing with OpenAI Whisper API: {audio_file_path} ...")
    try:
        with open(audio_file_path, "rb") as audio_file:
            transcription = stt_client.audio.transcriptions.create(
                model="whisper-1",
                file=audio_file
            )
        print("Whisper API transcription successful.")
        return f"Transcription for '{os.path.basename(audio_file_path)}':\n{transcription.text}"
    except Exception as e:
        return f"Error during Whisper API transcription for file {audio_file_path}: {e}"


@tool
def analyze_image(source: str, question: str) -> str:
    """
    Analyzes the visual content of an image to answer a specific question.
    This tool can examine an image from either a local file path (e.g., .png, .jpg) or a public URL. It understands the objects, scenes, text, and context within the image to provide a detailed answer.
    It returns a formatted string containing a header with the image source and the detailed answer to the question. On failure, it returns a string starting with 'Error:'.

    Args:
        source: The local file path OR the public URL of the image to analyze.
        question: The specific question to ask about the image's content.
    """
    if vlm_client is None:
        return "Error: The image analysis service is not configured. Please check system configuration."

    is_url = source.lower().startswith(('http://', 'https://'))
    source_display_name = source if is_url else os.path.basename(source)
    print(f"Analyzing image from '{source_display_name}'...")

    try:
        image_content_payload = {}
        if is_url:
            print("Input is a URL. Passing it directly to the analysis service.")
            image_content_payload = {"url": source}
        else:
            print("Input is a local file path. Encoding the image for analysis.")
            if not os.path.exists(source):
                return f"Error: Image file not found at path: {source}"
            image = Image.open(source)
            base64_image = _encode_pillow_image_to_base64(image)
            image_content_payload = {"url": base64_image}

        response = vlm_client.chat.completions.create(
            model=vlm_model,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": question},
                        {"type": "image_url", "image_url": image_content_payload},
                    ],
                }
            ],
            max_tokens=2000,
        )
        answer = response.choices[0].message.content
        print("Image analysis successful.")
        return f"Answer regarding '{source_display_name}':\n{answer}"

    except Exception as e:
        error_context = f"URL '{source}'" if is_url else f"image file '{source}'"
        return f"An error occurred during the analysis of {error_context}: {e}"


def _process_pdf_with_unstructured(file_path: str, page_numbers: Union[int, List[int], str],
                                   source_display_name: str) -> str:
    """Internal helper to process a local PDF file using UnstructuredPDFLoader."""
    try:
        print(f"Processing '{source_display_name}' with Unstructured loader...")
        loader = UnstructuredPDFLoader(file_path, mode="elements")
        all_docs = loader.load()

        if not all_docs:
            return f"The PDF from source '{source_display_name}' appears to be empty or unreadable."

        total_pages = max(doc.metadata.get('page_number', 0) for doc in all_docs) if all_docs else 0
        target_pages_1_based = []
        if isinstance(page_numbers, int):
            target_pages_1_based = [page_numbers]
        elif isinstance(page_numbers, list):
            target_pages_1_based = sorted(list(set(page_numbers)))
        elif page_numbers == 'all':
            num_to_process = min(total_pages, MAX_PAGES_TO_PROCESS)
            target_pages_1_based = list(range(1, num_to_process + 1))

        for p in target_pages_1_based:
            if not 1 <= p <= total_pages:
                return f"Error: Invalid page number {p} specified. The PDF has {total_pages} pages."
        if not target_pages_1_based:
            return "Error: No valid pages were specified for processing."

        content_by_page = {}
        for doc in all_docs:
            page_num = doc.metadata.get('page_number')
            if page_num in target_pages_1_based:
                if page_num not in content_by_page:
                    content_by_page[page_num] = []
                element_category = doc.metadata.get('category')
                if element_category == 'Image':
                    content_by_page[page_num].append(
                        '[Image detected, content cannot be read by this text-based tool.]')
                elif element_category == 'Table' and doc.metadata.get('text_as_html'):
                    content_by_page[page_num].append(
                        f"--- TABLE ---\n{doc.metadata['text_as_html']}\n--- END TABLE ---")
                else:
                    content_by_page[page_num].append(doc.page_content)

        if not content_by_page:
            return f"No content could be extracted from the specified page(s) of '{source_display_name}'. The file might be scanned or contain only images."

        full_text = [f"--- Page {pn} ---\n" + "\n\n".join(content).strip() for pn, content in
                     sorted(content_by_page.items())]
        final_output = "\n\n".join(full_text)

        if page_numbers == 'all' and total_pages > MAX_PAGES_TO_PROCESS:
            final_output += f"\n\n[Warning: The PDF has {total_pages} pages. Only the first {MAX_PAGES_TO_PROCESS} pages were processed.]"

        return final_output
    except Exception as e:
        return f"Error processing PDF '{source_display_name}' with Unstructured loader: {e}"


@tool
def read_pdf(source: str, page_numbers: Union[int, List[int], str] = 'all') -> str:
    """
    Extracts text and table data from specific pages of a PDF document, handling both local files and public URLs.
    IMPORTANT: This tool CANNOT read content from images or scanned pages (no OCR). It will insert a placeholder like '[Image detected...]' if it finds an image. For documents where visual elements (charts, diagrams) are crucial, use the 'ask_question_about_complex_document' tool instead. When 'all' pages are requested, processing is capped at the first 10 pages for efficiency.
    It returns a formatted string where content from each page is separated by a '--- Page X ---' header. Tables are returned as HTML within '--- TABLE ---' blocks, and a warning may be added for long documents. On failure, it returns a string starting with 'Error:'.

    Args:
        source: The local file path OR the public URL of the .pdf file.
        page_numbers: The page(s) to process. Can be an integer, a list of integers, or 'all'. Defaults to 'all'.
    """
    try:
        is_url = source.lower().startswith(('http://', 'https://'))
        source_display_name = source if is_url else os.path.basename(source)
        if is_url:
            response = requests.get(source)
            response.raise_for_status()
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=True) as temp_pdf:
                temp_pdf.write(response.content)
                temp_pdf.flush()
                return _process_pdf_with_unstructured(temp_pdf.name, page_numbers, source_display_name)
        else:
            if not os.path.exists(source):
                return f"Error: File not found at path: {source}"
            return _process_pdf_with_unstructured(source, page_numbers, source_display_name)
    except Exception as e:
        return f"An unexpected error occurred while handling '{source}': {e}"


def _filter_pdf_pages(original_path: str, page_numbers: Union[int, List[int]]) -> Optional[str]:
    reader = PdfReader(original_path)
    total_pages = len(reader.pages)
    target_indices = []
    pages_to_process = [page_numbers] if isinstance(page_numbers, int) else page_numbers

    for p_num in pages_to_process:
        if 1 <= p_num <= total_pages:
            target_indices.append(p_num - 1)
        else:
            print(f"Warning: Page number {p_num} is invalid for a PDF with {total_pages} pages. Ignoring.")

    if not target_indices:
        return None

    writer = PdfWriter()
    for index in target_indices:
        writer.add_page(reader.pages[index])

    temp_filtered_pdf = tempfile.NamedTemporaryFile(suffix=".pdf", delete=False)
    writer.write(temp_filtered_pdf.name)
    writer.close()
    print(f"Filtered PDF with pages {pages_to_process} created at: {temp_filtered_pdf.name}")
    return temp_filtered_pdf.name


def _process_document_with_vlm(file_path: str, question: str, page_numbers: Optional[Union[int, List[int], str]],
                               source_display_name: str) -> str:
    if doc_client is None:
        return "Error: Document Analysis client (doc_client) is not initialized."

    final_file_path = file_path
    temp_filtered_file_to_clean = None
    try:
        _, file_extension = os.path.splitext(file_path)
        if file_extension.lower() == '.pdf' and page_numbers not in (None, 'all'):
            print("Filtering PDF pages before sending to VLM...")
            filtered_path = _filter_pdf_pages(file_path, page_numbers)
            if not filtered_path:
                return "Error: No valid pages were selected for processing."
            final_file_path = filtered_path
            temp_filtered_file_to_clean = filtered_path

        print(f"Encoding '{os.path.basename(final_file_path)}' for VLM analysis...")
        with open(final_file_path, "rb") as f:
            base64_file = base64.b64encode(f.read()).decode('utf-8')

        mime_type = _get_file_mime_type(final_file_path)
        file_data_url = f"data:{mime_type};base64,{base64_file}"

        prompt = f"Analyze the attached document '{source_display_name}' and answer the question. Base your answer only on the document's content.\n\nQuestion: \"{question}\""
        if page_numbers and page_numbers != 'all':
            prompt += f"\nFocus on page(s)/slide(s): {page_numbers}."

        print("Sending request to the Document Analysis Model...")
        response = doc_client.chat.completions.create(
            model=doc_model,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": file_data_url}}
                    ]
                }
            ],
            max_tokens=3072
        )
        answer = response.choices[0].message.content
        return f"Answer from Document Analysis Model regarding '{source_display_name}':\n\n{answer}"

    except Exception as e:
        return f"An error occurred during VLM processing for '{source_display_name}': {e}"
    finally:
        if temp_filtered_file_to_clean and os.path.exists(temp_filtered_file_to_clean):
            os.remove(temp_filtered_file_to_clean)
            print(f"Cleaned up temporary filtered file: {temp_filtered_file_to_clean}")


@tool
def ask_question_about_complex_document(path_or_url: str, question: str,
                                        page_numbers: Optional[Union[int, List[int], str]] = None) -> str:
    """
    Answers specific questions about complex documents (PDF, DOCX, PPTX) by deeply analyzing their content, including text, tables, charts, and images.
    This is the primary tool to use when a simple text extraction from a 'read_*' tool is insufficient, for instance, if a 'read_*' tool returned an '[Image detected...]' warning, or if answering the question requires understanding visual elements like charts and graphs.
    It returns a detailed, context-aware answer based on the document's content, formatted with a header indicating the source. On failure, it returns a string starting with 'Error:'.

    Args:
        path_or_url: The local file path or public URL of the document.
        question: The specific question to ask about the document's content.
        page_numbers: Optional. The page(s) or slide(s) to focus on. Accepts 'all', a single number (e.g., 5), or a list (e.g., [1, 5, 10]).
    """
    is_url = path_or_url.lower().startswith(('http://', 'https://'))
    source_display_name = path_or_url if is_url else os.path.basename(path_or_url)
    _, file_extension = os.path.splitext(path_or_url)

    if not file_extension and is_url:
        try:
            with requests.head(path_or_url, allow_redirects=True, timeout=10) as h:
                content_type = h.headers.get('content-type', '').lower()
                if 'pdf' in content_type:
                    file_extension = '.pdf'
                elif 'vnd.openxmlformats-officedocument.wordprocessingml.document' in content_type:
                    file_extension = '.docx'
                elif 'vnd.openxmlformats-officedocument.presentationml.presentation' in content_type:
                    file_extension = '.pptx'
        except requests.RequestException:
            pass

    if file_extension.lower() not in ['.pdf', '.docx', '.pptx']:
        return f"Error: Unsupported file type '{file_extension}'. This tool only supports PDF, DOCX, and PPTX."

    try:
        if is_url:
            response = requests.get(path_or_url)
            response.raise_for_status()
            with tempfile.NamedTemporaryFile(suffix=file_extension, delete=True) as temp_doc:
                temp_doc.write(response.content)
                temp_doc.flush()
                return _process_document_with_vlm(temp_doc.name, question, page_numbers, source_display_name)
        else:
            if not os.path.exists(path_or_url):
                return f"Error: File not found at path: {path_or_url}"
            return _process_document_with_vlm(path_or_url, question, page_numbers, source_display_name)
    except Exception as e:
        return f"An unexpected error occurred: {e}"


def _extract_text_from_docx_file(file_path: str, source_display_name: str) -> str:
    """Internal helper to extract text and element metadata from a local .docx file using Unstructured."""
    try:
        print(f"Extracting elements from '{source_display_name}' using Unstructured loader...")
        loader = UnstructuredWordDocumentLoader(file_path, mode="elements")
        docs = loader.load()
        if not docs:
            return f"Successfully processed '{source_display_name}', but no content was extracted."

        content_parts = []
        for doc in docs:
            category = doc.metadata.get('category')
            if category == 'Image':
                content_parts.append('[Image detected, content cannot be read by this text-based tool.]')
            elif category == 'Table' and doc.metadata.get('text_as_html'):
                content_parts.append(f"--- TABLE ---\n{doc.metadata['text_as_html']}\n--- END TABLE ---")
            else:
                content_parts.append(doc.page_content)

        content = "\n\n".join(content_parts)
        return f"Successfully read content from {source_display_name}:\n{content}"
    except Exception as e:
        return f"An error occurred during text extraction from '{source_display_name}': {e}"


@tool
def read_docx(path_or_url: str) -> str:
    """
    Extracts all plain text from a Microsoft Word (.docx) document, from either a local file path or a public URL.
    Note: This tool extracts plain text only and cannot interpret images, charts, or complex layouts. For documents where visual elements are important, use the 'ask_question_about_complex_document' tool.
    It returns a formatted string containing a header with the source filename and all the extracted text. On failure, it returns a string starting with 'Error:'.

    Args:
        path_or_url: The local file path or the public URL of the .docx document.
    """
    is_url = path_or_url.lower().startswith(('http://', 'https://'))
    source_display_name = path_or_url if is_url else os.path.basename(path_or_url)
    try:
        if is_url:
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(path_or_url, headers=headers)
            response.raise_for_status()
            with tempfile.NamedTemporaryFile(suffix=".docx", delete=True) as temp_doc:
                temp_doc.write(response.content)
                temp_doc.flush()
                return _extract_text_from_docx_file(temp_doc.name, source_display_name)
        else:
            if not os.path.exists(path_or_url):
                return f"Error: File not found at path: {path_or_url}"
            return _extract_text_from_docx_file(path_or_url, source_display_name)
    except Exception as e:
        return f"An unexpected error occurred while handling '{source_display_name}': {e}"

@tool
def read_txt(path_or_url: str) -> str:
    """
    Extracts all plain text from a text (.txt) document, from either a local file path or a public URL.
    It returns a formatted string containing a header with the source filename and all the extracted text. On failure, it returns a string starting with 'Error:'.
    Args:
        path_or_url: The local file path or the public URL of the .txt document.
    """
    is_url = path_or_url.lower().startswith(('http://', 'https://'))
    source_display_name = path_or_url if is_url else os.path.basename(path_or_url)

    try:
        if is_url:
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(path_or_url, headers=headers)
            response.raise_for_status()
            text_content = response.text
            return f"Source: {source_display_name}\n\n{text_content}"
        else:
            if not os.path.exists(path_or_url):
                return f"Error: File not found at path: {path_or_url}"

            with open(path_or_url, 'r', encoding='utf-8', errors='ignore') as f:
                text_content = f.read()
            return f"Source: {source_display_name}\n\n{text_content}"

    except Exception as e:
        return f"An unexpected error occurred while handling '{source_display_name}': {e}"

@tool
def read_excel_data(source: str) -> str:
    """
    Reads all sheets from a Microsoft Excel file (.xlsx), extracting both the data content AND the background color of each cell.

    This tool is specifically designed for a Code Agent. It processes an Excel file from a local path or a public URL.
    For each sheet, it converts the data into a list of dictionaries (rows). The final output is a JSON string
    representing a dictionary where keys are sheet names.

    **IMPORTANT**: The structure for each cell is now a dictionary containing its value and color.
    - To access a value: `row['ColumnName']['value']`
    - To access a color: `row['ColumnName']['background_color']` (as an ARGB hex string, e.g., 'FFFF0000' for red)

    The agent can and should parse this JSON string directly in its code interpreter, for example:
    1. Parse the string: `import json; excel_data = json.loads(json_string_from_tool)`
    2. Create a DataFrame (for values only):
       `import pandas as pd;`
       `sheet1_data = excel_data['Sheet1']`
       `df_sheet1 = pd.DataFrame([{k: v['value'] for k, v in row.items()} for row in sheet1_data])`

    Returns a json string containing the data and styles from all sheets, or a string starting with 'Error:' on failure.

    Args:
        source (str): The local file path OR the public URL of the .xlsx file.
    """
    try:
        is_url = source.lower().startswith(('http://', 'https://'))
        source_display_name = source if is_url else os.path.basename(source)

        target = None
        if is_url:
            print(f"Downloading XLSX from URL: {source}")
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(source, headers=headers)
            response.raise_for_status()
            target = io.BytesIO(response.content)
        else:
            if not os.path.exists(source):
                return f"Error: File not found at local path: {source}"
            target = source

        print(f"Reading all sheets (with styles) from '{source_display_name}'...")
        workbook = openpyxl.load_workbook(target, data_only=True)

        all_sheets_data = {}
        for sheet in workbook:
            if sheet.max_row == 0:
                all_sheets_data[sheet.title] = []
                continue

            headers = [openpyxl.utils.get_column_letter(i) for i in range(1, sheet.max_column + 1)]

            sheet_rows_data = []
            for row in sheet.iter_rows(min_row=1):
                row_data = {}
                for header, cell in zip(headers, row):
                    value = cell.value
                    bg_color = "FFFFFFFF"

                    try:
                        if cell.fill.patternType == 'solid':
                            color = cell.fill.fgColor.rgb
                            if color and color != '00000000':
                                bg_color = color
                    except (AttributeError, TypeError):
                        pass

                    row_data[header] = {
                        "value": value,
                        "background_color": bg_color
                    }

                if any(cell_data['value'] is not None or cell_data['background_color'] != 'FFFFFFFF' for cell_data in
                       row_data.values()):
                    sheet_rows_data.append(row_data)

            all_sheets_data[sheet.title] = sheet_rows_data

        if not all_sheets_data:
            return f"The Excel file '{source_display_name}' was read successfully but it appears to be empty or contains no data/styled cells."

        final_json_string = json.dumps(all_sheets_data, indent=2, default=str)
        return final_json_string

    except Exception as e:
        return f"An error occurred while reading the Excel file '{source}': {e}"


def _process_pptx_with_unstructured(file_path: str, source_display_name: str) -> str:
    """Internal helper to process a local .pptx file, grouping content by slide."""
    try:
        print(f"Processing '{source_display_name}' with Unstructured loader...")
        loader = UnstructuredPowerPointLoader(file_path, mode="elements")
        docs = loader.load()
        if not docs:
            return f"Successfully processed '{source_display_name}', but no content was extracted."

        content_by_slide = {}
        for doc in docs:
            slide_num = doc.metadata.get('page_number', 'Unknown')
            if slide_num not in content_by_slide:
                content_by_slide[slide_num] = []

            category = doc.metadata.get('category')
            if category == 'Image':
                content_by_slide[slide_num].append('[Image detected, content cannot be read by this text-based tool.]')
            elif category == 'Table' and doc.metadata.get('text_as_html'):
                content_by_slide[slide_num].append(f"--- TABLE ---\n{doc.metadata['text_as_html']}\n--- END TABLE ---")
            else:
                content_by_slide[slide_num].append(doc.page_content)

        sorted_slides = sorted(content_by_slide.items(), key=lambda item: (isinstance(item[0], str), item[0]))
        full_text = [f"--- Slide {sn} ---\n" + "\n\n".join(c).strip() for sn, c in sorted_slides]
        content = "\n\n".join(full_text)
        return f"Successfully read content from {source_display_name}:\n{content}"
    except Exception as e:
        return f"An error occurred during content extraction from '{source_display_name}': {e}"


@tool
def read_pptx(source: str) -> str:
    """
    Extracts all text content (titles, body text, notes) from every slide of a Microsoft PowerPoint presentation (.pptx), from either a local path or a public URL.
    Note: This tool extracts text only and CANNOT interpret the visual content of slides, such as images, charts, or diagrams. For questions that require understanding these visual elements, use the 'ask_question_about_complex_document' tool.
    It returns a formatted string where content from each slide is separated by a header. On failure, it returns a string starting with 'Error:'.

    Args:
        source: The local file path OR the public URL of the .pptx file.
    """
    try:
        is_url = source.lower().startswith(('http://', 'https://'))
        source_display_name = source if is_url else os.path.basename(source)
        if is_url:
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(source, headers=headers)
            response.raise_for_status()
            with tempfile.NamedTemporaryFile(suffix=".pptx", delete=True) as temp_pptx:
                temp_pptx.write(response.content)
                temp_pptx.flush()
                return _process_pptx_with_unstructured(temp_pptx.name, source_display_name)
        else:
            if not os.path.exists(source):
                return f"Error: File not found at local path: {source}"
            return _process_pptx_with_unstructured(source, source_display_name)
    except Exception as e:
        return f"An unexpected error occurred while handling '{source}': {e}"


@tool
def unzip_file(zip_path: str, extract_to: str = None) -> str:
    """
    Unzips a .zip archive to a specified directory.
    It returns a string summarizing the result, including the list of extracted files. On failure, it returns a string starting with 'Error:'.

    Args:
        zip_path: The local path to the .zip file to be extracted.
        extract_to: Optional. The directory where contents should be extracted. If omitted, a new directory with the same name as the zip file (sans extension) is created in the same location.
    """
    if not os.path.exists(zip_path):
        return f"Error: The file {zip_path} does not exist."
    if extract_to is None:
        extract_to = os.path.splitext(zip_path)[0]
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(extract_to)
            extracted_files = zip_ref.namelist()
        return f"Successfully extracted {len(extracted_files)} items to '{extract_to}'. Files are: {extracted_files}"
    except Exception as e:
        return f"Error unzipping file {zip_path}: {e}"


@tool
def analyze_video(source: str, question: str) -> str:
    """
    Answers a specific question by analyzing the visual and audio content of a video from a local file or a public URL.
    This tool examines the video frame-by-frame and transcribes its audio to provide a comprehensive answer.
    Note: This is a powerful and resource-intensive tool, so analysis may take some time. Processing very large local files may fail due to size limits.
    It returns a detailed textual answer based on the analysis, prefixed with a header indicating the source. On failure, it returns a string starting with 'Error:'.

    Args:
        source: The local file path OR the public URL of the video to be analyzed.
        question: The specific question to ask about the video's content.
    """
    if not video_client:
        return "Error: The video analysis service is not configured. Please check system environment variables."

    is_url = source.lower().startswith(('http://', 'https://'))
    source_display_name = source if is_url else os.path.basename(source)

    try:
        video_data_url = ""

        if is_url:
            is_youtube_url = 'youtube.com' in source or 'youtu.be' in source

            if is_youtube_url:
                print(f"YouTube URL detected. Extracting direct video link using yt-dlp for '{source_display_name}'...")
                if not yt_dlp:
                    return "Error: yt-dlp library is required to process YouTube URLs. Please install it using 'pip install yt-dlp'."

                ydl_opts = {
                    'format': 'best[ext=mp4][height<=480]/best[ext=mp4]/best',
                    'quiet': True
                }
                with yt_dlp.YoutubeDL(ydl_opts) as ydl:
                    info = ydl.extract_info(source, download=False)
                    formats = info.get('formats', [info])
                    video_data_url = formats[0].get('url')
                    if not video_data_url:
                        return f"Error: Could not extract a direct video stream URL from '{source_display_name}' using yt-dlp."
                print("Direct video link extracted successfully.")

            else:
                video_data_url = source

            messages_payload = [{
                "role": "user",
                "content": [
                    {"type": "text", "text": f"Analyze this video and answer: {question}"},
                    {"type": "video_url", "video_url": {"url": video_data_url}}
                ]
            }]

        else:
            if not os.path.exists(source):
                return f"Error: Local video file not found at: {source}"
            print(f"Encoding local video file '{source_display_name}' to Base64...")
            with open(source, "rb") as video_file:
                base64_video = base64.b64encode(video_file.read()).decode('utf-8')

            video_data_url = f"data:video/mp4;base64,{base64_video}"
            messages_payload = [{
                "role": "user",
                "content": [
                    {"type": "text", "text": f"Analyze this video and answer: {question}"},
                    {"type": "image_url", "image_url": {"url": video_data_url}}
                ]
            }]

        print("Sending request to the video analysis model...")
        response = video_client.chat.completions.create(model=video_model, messages=messages_payload, max_tokens=2048)

        analysis_content = response.choices[0].message.content
        if not analysis_content.strip():
            return f"Analysis for video '{source_display_name}':\n\nThe model returned an empty response. This might happen if the video is too long, inaccessible, or the model failed to process it."

        return f"Analysis for video '{source_display_name}':\n\n{analysis_content}"

    except Exception as e:
        return f"An error occurred while analyzing the video '{source_display_name}': {e}"

class VisitWebpageTool(Tool):
    name = "visit_webpage"
    description = (
        "Visits a webpage at the given url and reads its content as a markdown string. Use this to browse webpages."
    )
    inputs = {
        "url": {
            "type": "string",
            "description": "The url of the webpage to visit.",
        }
    }
    output_type = "string"

    def __init__(self, max_output_length: int = 40000):
        super().__init__()
        self.max_output_length = max_output_length

    def _truncate_content(self, content: str, max_length: int) -> str:
        if len(content) <= max_length:
            return content
        return (
            content[: max_length // 2]
            + f"\n..._This content has been truncated to stay below {max_length} characters_...\n"
            + content[-max_length // 2 :]
        )

    def forward(self, url: str) -> str:
        try:
            import re
            import requests
            from markdownify import markdownify
            from requests.exceptions import RequestException
        except ImportError as e:
            raise ImportError(
                "You must install packages `markdownify` and `requests` to run this tool: for instance run `pip install markdownify requests`."
            ) from e
        try:
            headers = {
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
            }
            response = requests.get(url, headers=headers, timeout=20)

            response.raise_for_status()
            markdown_content = markdownify(response.text).strip()

            markdown_content = re.sub(r"\n{3,}", "\n\n", markdown_content)

            return self._truncate_content(markdown_content, self.max_output_length)

        except requests.exceptions.Timeout:
            return "The request timed out. Please try again later or check the URL."
        except RequestException as e:
            return f"Error fetching the webpage: {str(e)}"
        except Exception as e:
            return f"An unexpected error occurred: {str(e)}"

class WebSearchTool(Tool):
    name = "web_search"
    description = "Performs a web search for a query and returns a string of the top search results formatted as markdown with titles, links, and descriptions."
    inputs = {"query": {"type": "string", "description": "The search query to perform."}}
    output_type = "string"

    def __init__(self, max_results: int = 10, engine: str = "google"):
        super().__init__()
        self.max_results = max_results
        self.engine = engine

        if self.engine == "google":
            self.google_api_key = os.getenv("GOOGLE_API")
            self.google_cse_id = os.getenv("GOOGLE_ID")
            if not self.google_api_key or not self.google_cse_id:
                raise ValueError(
                    "For Google search, you must set the GOOGLE_API_KEY and GOOGLE_CSE_ID environment variables.")


    def forward(self, query: str) -> str:
        results = self.search(query)
        if len(results) == 0:
            raise Exception("No results found! Try a less restrictive/shorter query.")
        return self.parse_results(results)

    def search(self, query: str) -> list:
        if self.engine == "duckduckgo":
            return self.search_duckduckgo(query)
        elif self.engine == "bing":
            return self.search_bing(query)
        elif self.engine == "google":
            return self.search_google(query)
        else:
            raise ValueError(f"Unsupported engine: {self.engine}")

    def parse_results(self, results: list) -> str:
        return "## Search Results\n\n" + "\n\n".join(
            [f"[{result['title']}]({result['link']})\n{result['description']}" for result in results]
        )

    def search_google(self, query: str) -> list:
        url = "https://www.googleapis.com/customsearch/v1"
        params = {
            'key': self.google_api_key,
            'cx': self.google_cse_id,
            'q': query,
            'num': self.max_results
        }

        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()

        search_results = response.json()

        if 'error' in search_results:
            error_details = search_results['error']['message']
            raise Exception(f"Google API returned an error: {error_details}")

        items = search_results.get("items", [])

        results = [
            {
                "title": item.get("title"),
                "link": item.get("link"),
                "description": item.get("snippet")
            }
            for item in items
        ]
        return results

    def search_duckduckgo(self, query: str) -> list:
        params = {"q": query, "kl": "us-en"}

        response = requests.get(
            "https://lite.duckduckgo.com/lite/",
            params=params,
            headers={"User-Agent": "Mozilla/5.0"},
        )
        response.raise_for_status()
        parser = self._create_duckduckgo_parser()
        parser.feed(response.text)
        return parser.results[:self.max_results]

    def _create_duckduckgo_parser(self):
        class SimpleResultParser(HTMLParser):
            def __init__(self):
                super().__init__()
                self.results = []
                self.current = {}
                self.capture_title = False
                self.capture_description = False
                self.capture_link = False

            def handle_starttag(self, tag, attrs):
                attrs = dict(attrs)
                if tag == "a" and attrs.get("class") == "result-link":
                    self.capture_title = True
                elif tag == "td" and attrs.get("class") == "result-snippet":
                    self.capture_description = True
                elif tag == "span" and attrs.get("class") == "link-text":
                    self.capture_link = True

            def handle_endtag(self, tag):
                if tag == "a" and self.capture_title:
                    self.capture_title = False
                elif tag == "td" and self.capture_description:
                    self.capture_description = False
                elif tag == "span" and self.capture_link:
                    self.capture_link = False
                elif tag == "tr":
                    if {"title", "description", "link"} <= self.current.keys():
                        self.current["description"] = " ".join(self.current["description"])
                        self.results.append(self.current)
                        self.current = {}

            def handle_data(self, data):
                if self.capture_title:
                    self.current["title"] = data.strip()
                elif self.capture_description:
                    self.current.setdefault("description", [])
                    self.current["description"].append(data.strip())
                elif self.capture_link:
                    self.current["link"] = "https://" + data.strip()

        return SimpleResultParser()

    def search_bing(self, query: str) -> list:
        params = {"q": query, "format": "rss", "mkt": "en-US"}

        response = requests.get(
            "https://www.bing.com/search",
            params=params,
        )
        response.raise_for_status()
        root = ET.fromstring(response.text)
        items = root.findall(".//item")
        results = [
            {
                "title": item.findtext("title"),
                "link": item.findtext("link"),
                "description": item.findtext("description"),
            }
            for item in items[: self.max_results]
        ]
        return results

def get_correct_tools(requested_tools: List[str]) -> List[Tool]:

    print("Initializing toolset...")

    wikipedia_search_tool = WikipediaSearchTool(
        user_agent="new_tasks-Agent-Project (guodadipku@gmail.com)",
        content_type="text"
    )
    available_tools: Dict[str, Any] = {
        # "read_file": Tool.from_langchain(ReadFileTool()),
        "write_file": Tool.from_langchain(WriteFileTool()),
        "list_dir": Tool.from_langchain(ListDirectoryTool()),
        "visit_webpage": VisitWebpageTool(),
        "web_search": WebSearchTool(),
        "wiki_search": wikipedia_search_tool,
        "read_txt": read_txt,
        "read_pdf": read_pdf,
        "read_docx": read_docx,
        "read_pptx": read_pptx,
        "read_excel": read_excel_data,
        "speech_to_text": speech_to_text,
        "analyze_image": analyze_image,
        "analyze_video": analyze_video,
        "ask_document": ask_question_about_complex_document,
    }
    final_toolset = []
    for tool_name in requested_tools:
        tool = available_tools.get(tool_name)
        if tool:
            final_toolset.append(tool)
        else:
            print(f"Warning: Tool '{tool_name}' not found in available tools and will be ignored.")

    tool_names = [tool.name for tool in final_toolset]
    if len(tool_names) != len(set(tool_names)):
        print("Warning: Duplicate tool names found in the final toolset! This can cause unpredictable behavior.")
        import collections
        duplicates = [item for item, count in collections.Counter(tool_names).items() if count > 1]
        print(f"Duplicate names: {duplicates}")
    print(f"Toolset initialized with {len(final_toolset)} tools: {[tool.name for tool in final_toolset]}")
    return final_toolset